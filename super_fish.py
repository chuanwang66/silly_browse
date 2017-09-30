#-*- coding:utf-8 -*-
import os
from selenium import webdriver
import sys
import time

import openpyxl

account = ""
password = ""

class WorkSheet:
	def __init__(self, workbook_name, worksheet_name='sheet1'):
		self.workbook_name = workbook_name
		self.worksheet_name = worksheet_name
		self.loaded = False

	def create(self):
		self.wb = openpyxl.Workbook()
		self.ws = self.wb.get_active_sheet()
		self.ws.title = self.worksheet_name
		self.wb.save(self.workbook_name)
		self.wb.close()

	def remove(self):
		if self.exists():
			os.remove(self.workbook_name)

	def exists(self):
		return os.path.exists(self.workbook_name)

	def load(self):
		from openpyxl.reader.excel import load_workbook
		self.wb = load_workbook(self.workbook_name)

		if self.worksheet_name not in self.wb.get_sheet_names():
			print('worksheet[%s-%s] not existed'%(self.workbook_name, self.worksheet))
			return None

		self.ws = self.wb.get_sheet_by_name(self.worksheet_name)
		self.loaded = True

	def put(self, r, c, v):
		if not self.loaded:
			return
		self.ws.cell(row=r, column=c).value = v
		self.wb.save(self.workbook_name)

	def get(self, r, c):
		if not self.loaded:
			return None
		return self.ws.cell(row=r, column=c).value

	def close(self):
		self.wb.close()

if __name__ == '__main__':
	chrome_driver_path = os.path.join(os.getcwd(), 'chromedriver.exe')
	driver = webdriver.Chrome(chrome_driver_path)

	ws = WorkSheet(workbook_name='super_fish.xlsx', worksheet_name='live_info')
	if ws.exists():
		ws.remove()
	ws.create()
	ws.load()
	ws_row = 0

	#登录
	driver.get('http://open.talk-fun.com')
	driver.find_element_by_id("inputAccount").send_keys(account)
	driver.find_element_by_id("inputPassword").send_keys(password)
	driver.find_element_by_class_name('btn').click()

	while 1:
		try:
			link = driver.find_element_by_link_text(r'直播记录')
			link.click()
			break
		except:
			pass
		time.sleep(1.0)

	page = 0
	while 1:
		page += 1
		print('try page: %d'%(page))

		#1. 翻开新一页
		driver.get('http://open.talk-fun.com/backcms/index.php?action=live&pid=27&mod=yesterday&page=%d'%(page))

		table_body = None
		page_timelast = 0.0
		page_timeout = 10.0
		last_page = False
		while 1:
			if page_timelast >= page_timeout:
				break
			try:
				table_body = driver.find_element_by_xpath("//table[@class='tab_list live_hight_tab']/tbody")
				table_rows = table_body.find_elements_by_xpath(".//tr")
				if len(table_rows) == 0:
					last_page = True
				break
			except:
				time.sleep(1.0)
				page_timelast += 1.0
				continue
		if last_page:
			break

		#2. 分析一页中的直播记录列表
		stream_list = []
		for table_tr in table_body.find_elements_by_xpath(".//tr"):

			#直播ID
			try:
				stream_id = table_tr.find_element_by_xpath(".//td[@class='record_tb_std']/span").text.strip('\r\n\t ')
				room_name = table_tr.find_element_by_xpath(".//td[@class='record_tb_std re_name']/div[@class='rname']/span/em").text.strip('\r\n\t ')
				visitor_link = table_tr.find_element_by_xpath(".//td[@class='record_tb_std record_std']/span/a[@title='直播总访客数']").get_attribute('href')
				stream_list.append({'stream_id': str(stream_id), 'room_name': str(room_name), 'visitor_link': visitor_link})
			except:
				try:
					stream_list.append({'stream_id': str(stream_id), 'room_name': str(room_name), 'visitor_link': ''})
				except:
					print('visitor_link is not available: stream_id: ', stream_id, ', room_name: ', room_name)
					with open('stream_error.log', 'a+') as f:
						f.write('page: %d\t\t'%(page))
						f.write('stream_id: %s\t\t'%(stream_id))
						f.write('room_name: %s\n'%(room_name))


		#3. 对每一条记录再点进去看“最高同时在线人数”
		stream_peak_dict = {}
		for stream in stream_list:
			visitor_link = stream['visitor_link']
			if not visitor_link:
				stream_peak_dict[stream['stream_id']] = 0
			else:
				driver.get(visitor_link)
	
				while 1:
					try:
						driver.find_elements_by_xpath("//div[@class='box']")
						break
					except:
						time.sleep(1.0)
						continue

				for box in driver.find_elements_by_xpath("//div[@class='box']"):
					num = box.find_element_by_xpath(".//span[@class='num']").text.strip('\r\t\n ')
					caller = box.find_element_by_xpath(".//span[@class='caller']").text.strip('\r\t\n ')
					if '最高同时在线人数' in caller:
						try:
							stream_peak_dict[stream['stream_id']] = int(num)
						except:
							pass

		for stream in stream_list:
			try:
				ws_row += 1
				line = '%s\t\t%s\t\t%d'%(stream['stream_id'], stream['room_name'], stream_peak_dict[stream['stream_id']])
				print(line)
				ws.put(ws_row, 1, '%s'%(stream['stream_id']))
				ws.put(ws_row, 2, '%s'%(stream['room_name']))
				ws.put(ws_row, 3, '%d'%(stream_peak_dict[stream['stream_id']]))
			except Exception as e:
				print('error occurs: ', e)

	ws.close()
