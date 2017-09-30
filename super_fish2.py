#-*- coding:utf8 -*-
import io
from lxml import etree, html
import openpyxl
import os
import requests
import sys

class WorkSheet(object):
	def __init__(self, workbook_name, worksheet_name='sheet1'):
		self.workbook_name = workbook_name
		self.worksheet_name = worksheet_name
		self.loaded = False

	def create(self):
		self.wb = openpyxl.Workbook()
		self.ws = self.wb.get_active_sheet()
		self.ws.title = self.worksheet_name
		self.wb.save(self.workbook_name)
		self.wb.close()

	def remove(self):
		if self.exists():
			os.remove(self.workbook_name)

	def exists(self):
		return os.path.exists(self.workbook_name)

	def load(self):
		from openpyxl.reader.excel import load_workbook
		self.wb = load_workbook(self.workbook_name)

		if self.worksheet_name not in self.wb.get_sheet_names():
			print('worksheet[%s-%s] not existed'%(self.workbook_name, self.worksheet))
			return None

		self.ws = self.wb.get_sheet_by_name(self.worksheet_name)
		self.loaded = True

	def put(self, r, c, v):
		if not self.loaded:
			return
		self.ws.cell(row=r, column=c).value = v
		self.wb.save(self.workbook_name)

	def get(self, r, c):
		if not self.loaded:
			return None
		return self.ws.cell(row=r, column=c).value

	def close(self):
		self.wb.close()

def sizeof_fmt(num, suffix='B'):
    for unit in ['','K','M','G','T','P','E','Z']:
        if abs(num) < 1024.0:
            return "%3.1f%s%s" % (num, unit, suffix)
        num /= 1024.0
    return "%.1f%s%s" % (num, 'Y', suffix)

def check_connection(url):
    try:
        request = requests.head(url)
        if request.status_code == 200:
            print('Connection ok...')
            return True
        elif request.status_code == 301:
            print('Connection redirect')
            return True
        else:
            print('Error connecting')
            return False
    except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
        print('{} can not be reached, check your connection and retry later'.format(url))
        return False
    except (requests.exceptions.HTTPError, requests.exceptions.TooManyRedirects):
        print('There is an issue with the url, {}, confirm it, or retry later'.format(url))
        return False

raw_cookies = "Hm_lvt_37446e3c6b7cb9f5f065c0679a16722c=1493854771; Hm_lvt_9b54ef29e6bd23d6bfe22f68267014f9=1493854771; uid=open_100116; bid=0; roomid=552160; nickname=%E5%8A%A9%E6%95%99; t=1506678295; token=ca2dc3f6ccfb823f4ec80e9080a7e87a; access_token=wgjNkhzM0UzY2kTMxMTYiBjNxAjZjJWMzUGM1MzYmVDf8xXfiAjNxITN1IiOiUWbh5mciwCM6ISYiwCM6ICZpdmIsAjOiQWafV2cyV3bjJCLiIiOiIXY0FmdhJCLwojIyVGZuV2ZiwSM0IjM0UzNwUTM6ISZylGc4VmIsEDNygzN2YDM1EjOiUWbpR3ZlJnIs01W6Iic0RXYiwiIulWbkFmI6ISZs9mciwSO0MjN0ITM6ICZphnIsISO1UjN1xVOhJTN1xlI6ISZtFmbrNWauJCLwojIklmYiwiI2ETMwATMf5WZw9mI6ICZpVnIsAjNxITN1ojIklWbv9mciwCMyojIklGciwCMyojIkl2XyVmb0JXYwJye; PHPSESSID=jv5r9ohqgk5gdae58fkr3ap8p3"
def build_cookies():
	cookies = {}
	for line in raw_cookies.split(';'):
		key, value = line.split('=', 1)
		key, value = key.strip('\r\n\t '), value.strip('\r\n\t ')
		cookies[key] = value
	return cookies

if __name__ == "__main__":
	ws = WorkSheet(workbook_name='super_fish.xlsx', worksheet_name='live_info')
	if ws.exists():
		ws.remove()
	ws.create()
	ws.load()
	ws_row = 0

	#登录（手工完成，抓包得到cookies填入）
	cookies = build_cookies()
	
	# 翻页并抓取
	page = 0
	while 1:
		page += 1
		print('try page: %d'%(page))

		#1. 翻开新一页
		res = requests.get('http://open.talk-fun.com/backcms/index.php?action=live&pid=27&mod=yesterday&page=%d'%(page), cookies=cookies)
		html = etree.HTML(res.text)

		#sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')
		#print(res.text)

		#2. 分析一页中的直播记录列表
		last_page = False
		try:
			table_body = html.xpath("//table[@class='tab_list live_hight_tab']/tbody")[0]
			table_rows = table_body.xpath(".//tr")
			if len(table_rows) == 0:
				last_page = True
		except:
			pass
		if last_page:
			break

		stream_list = []
		for table_tr in table_body.xpath(".//tr"):
			try:
				stream_id = table_tr.xpath(".//td[@class='record_tb_std']/span")[0].text.strip('\r\n\t ')
				room_name = table_tr.xpath(".//td[@class='record_tb_std re_name']/div[@class='rname']/span/em")[0].text.strip('\r\n\t ')
				visitor_link = table_tr.xpath(".//td[@class='record_tb_std record_std']/span/a[@title='直播总访客数']")[0].get('href')
				visitor_link = "http://open.talk-fun.com/backcms/index.php%s"%(visitor_link)
				stream_list.append({'stream_id': str(stream_id), 'room_name': str(room_name), 'visitor_link': visitor_link})
			except:
				try:
					stream_list.append({'stream_id': str(stream_id), 'room_name': str(room_name), 'visitor_link': ''})
				except:
					print('visitor_link is not available: stream_id: ', stream_id, ', room_name: ', room_name)
					with open('stream_error.log', 'a+') as f:
						f.write('page: %d\t\t'%(page))
						f.write('stream_id: %s\t\t'%(stream_id))
						f.write('room_name: %s\n'%(room_name))

		#3. 对每一条记录再点进去看“最高同时在线人数”
		stream_peak_dict = {}
		for stream in stream_list:
			visitor_link = stream['visitor_link']
			if not visitor_link:
				stream_peak_dict[stream['stream_id']] = 0
			else:
				res = requests.get(visitor_link, cookies=cookies)
				html = etree.HTML(res.text)

				for box in html.xpath("//div[@class='box']"):
					num = box.xpath(".//span[@class='num']")[0].text.strip('\r\n\t ')
					caller = box.xpath(".//span[@class='caller']")[0].text.strip('\r\n\t ')
					if '最高同时在线人数' in caller:
						try:
							stream_peak_dict[stream['stream_id']] = int(num)
						except:
							pass

		#4. 写文本
		for stream in stream_list:
			try:
				ws_row += 1
				line = '%s\t\t%s\t\t%d'%(stream['stream_id'], stream['room_name'], stream_peak_dict[stream['stream_id']])
				print(line)
				ws.put(ws_row, 1, '%s'%(stream['stream_id']))
				ws.put(ws_row, 2, '%s'%(stream['room_name']))
				ws.put(ws_row, 3, '%d'%(stream_peak_dict[stream['stream_id']]))
			except Exception as e:
				print('error occurs: ', e)

	ws.close()