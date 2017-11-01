#-*- coding:utf8 -*-
import atexit
import concurrent.futures
import io
from lxml import etree, html
import openpyxl
import os
import requests
import grequests
import sys
import threading

account = u""
password = u""

# get请求1个url
# max_retries: 若第1次请求失败，再尝试max_retries次
# verbose: 输出提示信息
# 返回: 成功则返回requests.Response; 失败则返回None
#
# 注: 本函数阻塞等待
# 注：本函数作用在于“可以重试”
def requests_get_url(url, cookies={}, max_retries=3, verbose=False):
    r = None

    num_retries = 0
    while num_retries <= max_retries:
        try:
            r = requests.get(url, cookies=cookies, timeout=10.000)
            break
        except requests.exceptions.ConnectTimeout:
            if verbose: print('connect timeout')
            num_retries += 1
        except requests.exceptions.ConnectionError:
            if verbose: print('connect error')
            num_retries += 1
        except requests.exceptions.Timeout:
            if verbose: print('timeout')
            num_retries += 1
        except requests.exceptions.TooManyRedirects:
            if verbose: print('too many redirects')
            num_retries += 1
        except requests.exceptions.RequestException as e:
            if verbose: print(e)
            num_retries += 1
        if verbose: print('retry: %d'%(num_retries))
    return r


# get批量请求url_list
# url_list: 请求url列表
# verbose: 输出提示信息
# 返回: ok_result(url -> requests.models.Response), fail_result(url -> Exception)
#
# 注: 本函数仍阻塞等待
# 注：本函数作用在于“可以并发请求多个url”
def grequests_get_urls(url_list, cookies={}, max_retries=3, verbose=False):
    ok_result = {}
    fail_result = {}

    def response_hook(r, **kwargs):
        ok_result[r.url] = r
        return r

    def exception_handler(r, exception):
        fail_result[r.url] = exception
        return exception

    num_retries = 0
    while num_retries <= max_retries:
        reqs = [grequests.get(url, cookies=cookies, hooks={'response': [response_hook]}, timeout=5.000) for url in url_list]
        reps = grequests.map(reqs, size=100, exception_handler=exception_handler)

        if not fail_result:
            break
        else:
            num_retries += 1

            url_list = []
            for url in fail_result.keys():
                url_list.append(url)

            if verbose:
                print('retry: %d'%(num_retries))
                for url in url_list:
                    print('\t%s'%(url))

    return ok_result, fail_result

def sizeof_fmt(num, suffix='B'):
    for unit in ['','K','M','G','T','P','E','Z']:
        if abs(num) < 1024.0:
            return "%3.1f%s%s" % (num, unit, suffix)
        num /= 1024.0
    return "%.1f%s%s" % (num, 'Y', suffix)

def build_cookies(raw_cookies):
    cookies = {}
    for line in raw_cookies.split(';'):
        key, value = line.split('=', 1)
        key, value = key.strip('\r\n\t '), value.strip('\r\n\t ')
        cookies[key] = value
    return cookies

class WorkSheet(object):
    def __init__(self, workbook_name, worksheet_name='sheet1'):
        self.workbook_name = workbook_name
        self.worksheet_name = worksheet_name
        self.loaded = False

    def create(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.get_active_sheet()
        self.ws.title = self.worksheet_name
        self.wb.save(self.workbook_name)
        self.wb.close()

    def remove(self):
        if self.exists():
            os.remove(self.workbook_name)

    def exists(self):
        return os.path.exists(self.workbook_name)

    def load(self):
        from openpyxl.reader.excel import load_workbook
        self.wb = load_workbook(self.workbook_name)

        if self.worksheet_name not in self.wb.get_sheet_names():
            print('worksheet[%s-%s] not existed'%(self.workbook_name, self.worksheet))
            return None

        self.ws = self.wb.get_sheet_by_name(self.worksheet_name)
        self.loaded = True

    def put(self, r, c, v):
        if not self.loaded:
            return
        self.ws.cell(row=r, column=c).value = v
        self.wb.save(self.workbook_name)

    def get(self, r, c):
        if not self.loaded:
            return None
        return self.ws.cell(row=r, column=c).value

    def close(self):
        self.wb.close()

_requests_executor = None

def exit_callback():
    _requests_executor.shutdown()

atexit.register(exit_callback)

def on_cookie_response(future):

    r = future.result()

    ws = WorkSheet(workbook_name='super_fish.xlsx', worksheet_name='live_info')
    if ws.exists():
        ws.remove()
    ws.create()
    ws.load()
    ws_row = 0

    raw_cookies = None
    for key in r.headers.keys():
        if 'Set-Cookie' in key:
            raw_cookies = r.headers[key]
    if not raw_cookies:
        print('get cookies failed')
        import sys
        sys.exit(1)
    else:
        print('get cookies ok: %s'%raw_cookies)

    cookies = build_cookies(raw_cookies)
    
    # 翻页并抓取
    stream_list = []
    page = 0
    while 1:
        page += 1
        print('try page: %d'%(page))

        #1. 翻开新一页
        res = requests_get_url('http://open.talk-fun.com/backcms/index.php?action=live&pid=27&mod=yesterday&page=%d'%(page), cookies=cookies, max_retries=3)
        #res = requests.get('http://open.talk-fun.com/backcms/index.php?action=live&pid=27&mod=yesterday&page=%d'%(page), cookies=cookies)
        html = etree.HTML(res.text)

        #sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')
        #print(res.text)

        #2. 分析一页中的直播记录列表
        last_page = False
        try:
            table_body = html.xpath("//table[@class='tab_list live_hight_tab']/tbody")[0]
            table_rows = table_body.xpath(".//tr")
            if len(table_rows) == 0:
                last_page = True
        except:
            pass
        if last_page:
            break

        for table_tr in table_body.xpath(".//tr"):
            try:
                stream_id = table_tr.xpath(".//td[@class='record_tb_std']/span")[0].text.strip('\r\n\t ')
                room_name = table_tr.xpath(".//td[@class='record_tb_std re_name']/div[@class='rname']/span/em")[0].text.strip('\r\n\t ')
                visitor_link = table_tr.xpath(".//td[@class='record_tb_std record_std']/span/a[@title='直播总访客数']")[0].get('href')
                visitor_link = "http://open.talk-fun.com/backcms/index.php%s"%(visitor_link)
                stream_list.append({'stream_id': str(stream_id), 'room_name': str(room_name), 'visitor_link': visitor_link})
            except:
                try:
                    stream_list.append({'stream_id': str(stream_id), 'room_name': str(room_name), 'visitor_link': ''})
                except:
                    print('visitor_link is not available: stream_id: ', stream_id, ', room_name: ', room_name)
                    with open('stream_error.log', 'a+') as f:
                        f.write('page: %d\t\t'%(page))
                        f.write('stream_id: %s\t\t'%(stream_id))
                        f.write('room_name: %s\n'%(room_name))

    if stream_list:
        #3. 并发请求: 对每一条记录点击去看“最高同时在线人数”
        stream_peak_dict = {}
        link_stream_dict = {}
        visitor_links = []
        for stream in stream_list:
            visitor_link = stream['visitor_link']
            if not visitor_link:
                stream_peak_dict[stream['stream_id']] = 0
            else:
                visitor_links.append(visitor_link)
                link_stream_dict[visitor_link] = stream['stream_id']

        ok_result, fail_result = grequests_get_urls(visitor_links, cookies=cookies, max_retries=3)
        if ok_result:
            for url in ok_result.keys():
                html = etree.HTML(ok_result[url].text)

                for box in html.xpath("//div[@class='box']"):
                    num = box.xpath(".//span[@class='num']")[0].text.strip('\r\n\t ')
                    caller = box.xpath(".//span[@class='caller']")[0].text.strip('\r\n\t ')
                    if '最高同时在线人数' in caller:
                        try:
                            stream_peak_dict[link_stream_dict[url]] = int(num)
                        except:
                            pass
        if fail_result:
            print(fail_result)

        #4. 写文本
        for stream in stream_list:
            try:
                ws_row += 1
                line = '%s\t\t%s\t\t%d'%(stream['stream_id'], stream['room_name'], stream_peak_dict[stream['stream_id']])
                print(line)
                ws.put(ws_row, 1, '%s'%(stream['stream_id']))
                ws.put(ws_row, 2, '%s'%(stream['room_name']))
                ws.put(ws_row, 3, '%d'%(stream_peak_dict[stream['stream_id']]))
            except Exception as e:
                print('error occurs: ', e)

    ws.close()

if __name__ == "__main__":

    _requests_executor = concurrent.futures.ThreadPoolExecutor(max_workers=5)

    ##登录, 获取cookies

    #r = requests.post("http://open.talk-fun.com/backcms/index.php?action=index&sub=login", data=post_data, headers=post_headers)
    future = _requests_executor.submit(requests.post, 
        "http://open.talk-fun.com/backcms/index.php?action=index&sub=login", 
        data={u"partner":u"", u"name":account, u"password":password},                   #post-data
        headers={"content-type": "application/x-www-form-urlencoded; charset=UTF-8"}    #post-headers
    )
    future.add_done_callback(on_cookie_response)