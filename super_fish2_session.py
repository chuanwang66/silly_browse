#-*- coding:utf8 -*-
import io
from lxml import etree, html
import openpyxl
import os
import requests
import grequests
import sys
import threading

account = u""
password = u""

###################################################################################### 测试(begin)
url_list = [
	"https://github.com/chuanwang66/silly_player_x",
	"https://github.com/chuanwang66/silly_browse",
	"https://github.com/chuanwang66/easerpc_x",
	"https://github.com/chuanwang66/kcp",
	"https://github.com/chuanwang66/test_opencv",
	"https://github.com/chuanwang66/C-Thread-Pool",
	"https://github.com/chuanwang66/ngx_stream_upstream_check_module",
	"https://github.com/chuanwang66/AsyncNet",
	"https://github.com/chuanwang66/Spider",
	"https://github.com/chuanwang66/AudioStreaming",
	"https://github3.com/chuanwang66/fake"
]

def test_grequests():
	"""
		grequests.map 和 grequests.imap 的优势在于并发执行多个网络请求；但仍是阻塞等待
	"""
	result = {}

	def response_hook(r, **kwargs):
		result[r.url] = True
		print('%s ==> %d'%(r.url, len(r.content)))
		return r

	def response_callback(r, **kwargs):
		print('%s ==> %d'%(r.url, len(r.content)))
		return r

	def exception_handler(r, exception):
		print('%s ==> %s'%(r.url, exception))
		return exception

	#map    并发执行，阻塞等待 (立即执行)
	#imap   并发执行，阻塞等待 (返回生成器，使用时开始执行)
	#send	并发执行，异步不等待
	map_imap_pool = 3
	if map_imap_pool == 1:
		#grequests.map: waits for all requests to finish, then returns the ordered data
		reqs = [grequests.get(url, hooks={'response': [response_hook]}, timeout=5.000) for url in url_list]
		reps = grequests.map(reqs, size=5, exception_handler=exception_handler)

		print('here')
	elif map_imap_pool == 2:
		#grequests.imap: is similar to map but returns a generator
		reqs = [grequests.get(url, callback=response_callback, timeout=5.000) for url in url_list]
		for r in grequests.imap(reqs, size=10):
			pass

		print('here')
	elif map_imap_pool == 3:
		#但这种方式不能设置exception_handler
		for url in url_list:
			req = grequests.get(url, hooks=dict(response=response_hook), timeout=5.000)
			job = grequests.send(req, grequests.Pool(10))

		print('here')

		import time
		time.sleep(20.0)
	else:
		pass
###################################################################################### 测试(end)

# get请求1个url
# max_retries: 若第1次请求失败，再尝试max_retries次
# verbose: 输出提示信息
# 返回: 成功则返回requests.Response; 失败则返回None
#
# 注: 本函数阻塞等待
# 注：本函数作用在于“可以重试”
def requests_get_url(url, cookies={}, max_retries=3, verbose=False):
	r = None

	num_retries = 0
	while num_retries <= max_retries:
		try:
			r = requests.get(url, cookies=cookies, timeout=10.000)
			break
		except requests.exceptions.ConnectTimeout:
			if verbose: print('connect timeout')
			num_retries += 1
		except requests.exceptions.ConnectionError:
			if verbose: print('connect error')
			num_retries += 1
		except requests.exceptions.Timeout:
			if verbose: print('timeout')
			num_retries += 1
		except requests.exceptions.TooManyRedirects:
			if verbose: print('too many redirects')
			num_retries += 1
		except requests.exceptions.RequestException as e:
			if verbose: print(e)
			num_retries += 1
		if verbose: print('retry: %d'%(num_retries))
	return r

def session_get_url(session, url, max_retries=3, verbose=False):
	r = None

	num_retries = 0
	while num_retries <= max_retries:
		try:
			r = session.get(url, timeout=10.000)
			break
		except requests.exceptions.ConnectTimeout:
			if verbose: print('connect timeout')
			num_retries += 1
		except requests.exceptions.ConnectionError:
			if verbose: print('connect error')
			num_retries += 1
		except requests.exceptions.Timeout:
			if verbose: print('timeout')
			num_retries += 1
		except requests.exceptions.TooManyRedirects:
			if verbose: print('too many redirects')
			num_retries += 1
		except requests.exceptions.RequestException as e:
			if verbose: print(e)
			num_retries += 1
		if verbose: print('retry: %d'%(num_retries))
	return r

# get批量请求url_list
# url_list: 请求url列表
# verbose: 输出提示信息
# 返回: ok_result(url -> requests.models.Response), fail_result(url -> Exception)
#
# 注: 本函数仍阻塞等待
# 注：本函数作用在于“可以并发请求多个url”
def grequests_get_urls(url_list, cookies={}, max_retries=3, verbose=False):
	ok_result = {}
	fail_result = {}

	def response_hook(r, **kwargs):
		ok_result[r.url] = r
		return r

	def exception_handler(r, exception):
		fail_result[r.url] = exception
		return exception

	num_retries = 0
	while num_retries <= max_retries:
		reqs = [grequests.get(url, cookies=cookies, hooks={'response': [response_hook]}, timeout=5.000) for url in url_list]
		reps = grequests.map(reqs, size=20, exception_handler=exception_handler)

		if not fail_result:
			break
		else:
			num_retries += 1

			url_list = []
			for url in fail_result.keys():
				url_list.append(url)

			if verbose:
				print('retry: %d'%(num_retries))
				for url in url_list:
					print('\t%s'%(url))

	return ok_result, fail_result

def session_get_urls(session, url_list, max_retries=3, verbose=False):
	ok_result = {}
	fail_result = {}

	def response_hook(r, **kwargs):
		ok_result[r.url] = r
		return r

	def exception_handler(r, exception):
		fail_result[r.url] = exception
		return exception

	num_retries = 0
	while num_retries <= max_retries:
		reqs = [session.get(url, hooks={'response': [response_hook]}, timeout=5.000) for url in url_list]
		reps = grequests.map(reqs, size=20, exception_handler=exception_handler)

		if not fail_result:
			break
		else:
			num_retries += 1

			url_list = []
			for url in fail_result.keys():
				url_list.append(url)

			if verbose:
				print('retry: %d'%(num_retries))
				for url in url_list:
					print('\t%s'%(url))

	return ok_result, fail_result

def check_connection(url):
    try:
        request = requests.head(url)
        if request.status_code == 200:
            print('Connection ok...')
            return True
        elif request.status_code == 301:
            print('Connection redirect')
            return True
        else:
            print('Error connecting')
            return False
    except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
        print('{} can not be reached, check your connection and retry later'.format(url))
        return False
    except (requests.exceptions.HTTPError, requests.exceptions.TooManyRedirects):
        print('There is an issue with the url, {}, confirm it, or retry later'.format(url))
        return False

def sizeof_fmt(num, suffix='B'):
    for unit in ['','K','M','G','T','P','E','Z']:
        if abs(num) < 1024.0:
            return "%3.1f%s%s" % (num, unit, suffix)
        num /= 1024.0
    return "%.1f%s%s" % (num, 'Y', suffix)

def build_cookies(raw_cookies):
	cookies = {}
	for line in raw_cookies.split(';'):
		key, value = line.split('=', 1)
		key, value = key.strip('\r\n\t '), value.strip('\r\n\t ')
		cookies[key] = value
	return cookies

class WorkSheet(object):
	def __init__(self, workbook_name, worksheet_name='sheet1'):
		self.workbook_name = workbook_name
		self.worksheet_name = worksheet_name
		self.loaded = False

	def create(self):
		self.wb = openpyxl.Workbook()
		self.ws = self.wb.get_active_sheet()
		self.ws.title = self.worksheet_name
		self.wb.save(self.workbook_name)
		self.wb.close()

	def remove(self):
		if self.exists():
			os.remove(self.workbook_name)

	def exists(self):
		return os.path.exists(self.workbook_name)

	def load(self):
		from openpyxl.reader.excel import load_workbook
		self.wb = load_workbook(self.workbook_name)

		if self.worksheet_name not in self.wb.get_sheet_names():
			print('worksheet[%s-%s] not existed'%(self.workbook_name, self.worksheet))
			return None

		self.ws = self.wb.get_sheet_by_name(self.worksheet_name)
		self.loaded = True

	def put(self, r, c, v):
		if not self.loaded:
			return
		self.ws.cell(row=r, column=c).value = v
		self.wb.save(self.workbook_name)

	def get(self, r, c):
		if not self.loaded:
			return None
		return self.ws.cell(row=r, column=c).value

	def close(self):
		self.wb.close()

if __name__ == "__main__":
	#test_grequests()
	#test_requests_get_url()
	#test_grequests_get_urls()

	ws = WorkSheet(workbook_name='super_fish.xlsx', worksheet_name='live_info')
	if ws.exists():
		ws.remove()
	ws.create()
	ws.load()
	ws_row = 0

	##登录, 获取cookies
	post_headers = {"content-type": "application/x-www-form-urlencoded; charset=UTF-8"}
	post_data = {u"partner":u"", u"name":account, u"password":password}
	r = requests.post("http://open.talk-fun.com/backcms/index.php?action=index&sub=login", data=post_data, headers=post_headers)

	raw_cookies = None
	for key in r.headers.keys():
		if 'Set-Cookie' in key:
			raw_cookies = r.headers[key]
	if not raw_cookies:
		print('get cookies failed')
		import sys
		sys.exit(1)
	else:
		print('get cookies ok: %s'%raw_cookies)

	cookies = build_cookies(raw_cookies)
	session = requests.Session()
	requests.utils.add_dict_to_cookiejar(session.cookies, cookies)
	
	# 翻页并抓取
	page = 0
	while 1:
		page += 1
		print('try page: %d'%(page))

		#1. 翻开新一页
		res = session_get_url(session, 'http://open.talk-fun.com/backcms/index.php?action=live&pid=27&mod=yesterday&page=%d'%(page), max_retries=3)
		#res = requests.get('http://open.talk-fun.com/backcms/index.php?action=live&pid=27&mod=yesterday&page=%d'%(page), cookies=cookies)
		html = etree.HTML(res.text)

		#sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')
		#print(res.text)

		#2. 分析一页中的直播记录列表
		last_page = False
		try:
			table_body = html.xpath("//table[@class='tab_list live_hight_tab']/tbody")[0]
			table_rows = table_body.xpath(".//tr")
			if len(table_rows) == 0:
				last_page = True
		except:
			pass
		if last_page:
			break

		stream_list = []
		for table_tr in table_body.xpath(".//tr"):
			try:
				stream_id = table_tr.xpath(".//td[@class='record_tb_std']/span")[0].text.strip('\r\n\t ')
				room_name = table_tr.xpath(".//td[@class='record_tb_std re_name']/div[@class='rname']/span/em")[0].text.strip('\r\n\t ')
				visitor_link = table_tr.xpath(".//td[@class='record_tb_std record_std']/span/a[@title='直播总访客数']")[0].get('href')
				visitor_link = "http://open.talk-fun.com/backcms/index.php%s"%(visitor_link)
				stream_list.append({'stream_id': str(stream_id), 'room_name': str(room_name), 'visitor_link': visitor_link})
			except:
				try:
					stream_list.append({'stream_id': str(stream_id), 'room_name': str(room_name), 'visitor_link': ''})
				except:
					print('visitor_link is not available: stream_id: ', stream_id, ', room_name: ', room_name)
					with open('stream_error.log', 'a+') as f:
						f.write('page: %d\t\t'%(page))
						f.write('stream_id: %s\t\t'%(stream_id))
						f.write('room_name: %s\n'%(room_name))

		#3. 对每一条记录再点进去看“最高同时在线人数”
		stream_peak_dict = {}
		for stream in stream_list:
			visitor_link = stream['visitor_link']
			if not visitor_link:
				stream_peak_dict[stream['stream_id']] = 0
			else:
				res = session_get_url(session, visitor_link, max_retries=3)
				#res = requests.get(visitor_link, cookies=cookies)
				html = etree.HTML(res.text)

				for box in html.xpath("//div[@class='box']"):
					num = box.xpath(".//span[@class='num']")[0].text.strip('\r\n\t ')
					caller = box.xpath(".//span[@class='caller']")[0].text.strip('\r\n\t ')
					if '最高同时在线人数' in caller:
						try:
							stream_peak_dict[stream['stream_id']] = int(num)
						except:
							pass

		#4. 写文本
		for stream in stream_list:
			try:
				ws_row += 1
				line = '%s\t\t%s\t\t%d'%(stream['stream_id'], stream['room_name'], stream_peak_dict[stream['stream_id']])
				print(line)
				ws.put(ws_row, 1, '%s'%(stream['stream_id']))
				ws.put(ws_row, 2, '%s'%(stream['room_name']))
				ws.put(ws_row, 3, '%d'%(stream_peak_dict[stream['stream_id']]))
			except Exception as e:
				print('error occurs: ', e)

	ws.close()